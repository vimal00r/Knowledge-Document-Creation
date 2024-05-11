import fitz
import pandas as pd
import numpy as np
import glob
import requests
import xlsxwriter
import pdfplumber
import openpyxl
import re
import os
import xlsxwriter
import PyPDF2

from bs4 import BeautifulSoup
from img2table.document import PDF
from img2table.ocr import TesseractOCR
from IPython.display import display_html
from bs4 import BeautifulSoup
from fitz.fitz import TEXT_ALIGN_LEFT


def get_tables(text, name):
    soup = BeautifulSoup(text, 'html.parser')  # Specify the parser explicitly
    tables = soup.find_all('table')  # Find all tables in the HTML
    for table_index, table in enumerate(tables):
        header = []
        rows = []
        for i, row in enumerate(table.find_all('tr')):
            cells = [cell.text.strip().replace('\n', '') for cell in row.find_all(['th', 'td'])]
            if i == 0:
                header = cells
            else:
                rows.append(cells)
        # Create a separate XLSX file for each table
        workbook = xlsxwriter.Workbook(name)
        worksheet = workbook.add_worksheet()
        worksheet.write_row(0, 0, header)
        for i, row in enumerate(rows):
            worksheet.write_row(i + 1, 0, row)
        workbook.close()


def get_all_pages_text(pdf_path):
    arr=[]
    def not_within_bboxes(obj):
        def obj_in_bbox(_bbox):
            v_mid = (obj["top"] + obj["bottom"]) / 2
            h_mid = (obj["x0"] + obj["x1"]) / 2
            x0, top, x1, bottom = _bbox
            return (h_mid >= x0) and (h_mid < x1) and (v_mid >= top) and (v_mid < bottom)
        return not any(obj_in_bbox(__bbox) for __bbox in bboxes)
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            try:
                bboxes = [
                    table.bbox
                    for table in page.find_tables(
                        table_settings={"vertical_strategy": "explicit",
                                        "horizontal_strategy": "explicit",
                                        "explicit_vertical_lines": page.curves + page.edges,
                                        "explicit_horizontal_lines": page.curves + page.edges,
                                        }
                    )
                ]
            except Exception as e:
                print(f"An error occurred: Image Appeared")
                continue
            arr.append(page.filter(not_within_bboxes).extract_text())
    return arr


def get_all_content(pdf_path):
    ocr=TesseractOCR(lang='eng')
    pdf=PDF(src=pdf_path)
    extracted_tables = pdf.extract_tables(ocr=ocr,
                                      implicit_rows=False,
                                      borderless_tables=False,
                                      min_confidence=50)
    #Saving html table to xlsx format
    n = 1
    for page, tables in extracted_tables.items():    
        for idx, table in enumerate(tables):
            z=table.html_repr()
            get_tables(z,f'./Table_to_text/tables/Table {n}.xlsx')
            n += 1
            # get_tables(z,f'./Table_to_text/tables/Page-{page+1} Table-{idx+1}.xlsx')



# Function to convert XLSX to TXT
def xlsx_to_txt(xlsx_file, txt_file):
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        with open(txt_file, 'w', encoding='utf-8') as txt_file:
            # Iterate through rows in the sheet and write cell values to the TXT file
            for row in sheet.iter_rows():
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                row_text = '\t'.join(row_data)  # Separate cell values by tabs
                txt_file.write(row_text + '\n')  # Write the row to the TXT file
    except Exception as e:
        print(f"Error: {e}")



def pdf_redact(doc,page_num,rectangle,text):
    doc[page_num].add_redact_annot(rectangle)
    doc[page_num].apply_redactions()
    tw = fitz.TextWriter(doc[page_num].rect)
    tw.fill_textbox(rectangle,text)
    tw.write_text(doc[page_num])



def read(txt_file_path):
    with open(txt_file_path, "r",encoding='utf-8') as file:
            file_contents = file.read()
    return file_contents


folder_path = './Table_to_text/tables/'

table_id_dict = {}

def table_to_text(pdf_path):

        get_all_content(pdf_path)

        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx'):
                xlsx_file = os.path.join(folder_path, filename)
                txt_file = os.path.join('./Table_to_text/tables_data/', filename.replace('.xlsx', '.txt'))
                xlsx_to_txt(xlsx_file, txt_file)
                with open(txt_file, "rb") as file:
                    file_contents = file.read()
                table_id_dict[filename.split(".")[0]] = file_contents

        doc=fitz.open(pdf_path)

        with pdfplumber.open(pdf_path) as pdf:
            n = 1
            for num,page in enumerate(pdf.pages): 
                k = [
                    table.bbox
                    for table in page.find_tables(
                        table_settings={"vertical_strategy": "lines",
                                        "horizontal_strategy": "lines",
                                        # "explicit_vertical_lines": [100.0, 200.0, 300.0] ,
                                        # "explicit_horizontal_lines": [100.0, 200.0, 300.0] ,
                                        }
                    )
                ]
                # k=page.find_tables(table_settings={"vertical_strategy": "explicit",
                #                                     "horizontal_strategy": "explicit",
                #                                     "explicit_vertical_lines": page.curves + page.edges,
                #                                     "explicit_horizontal_lines": page.curves + page.edges,
                #                                     })
                
                for table,z in enumerate(k):
                    pdf_redact(doc,num,fitz.Rect(z),f"Table {n}")
                    n+=1
                    # pdf_redact(doc,num,fitz.Rect(z),f"Page-{num+1} Table-{table+1}")
                    # pdf_redact(doc,num,fitz.Rect(z.bbox),str(read("./Table_to_text/tables_data/"+f"Page-{num+1} Table-{table+1}.txt")))
   
        doc.save('./table_to_text_output.pdf',deflate=True)        
        doc.close()
        print("Table_id_dict")
        print(table_id_dict)
        print("********************************************************************")
        
        return table_id_dict

#         print(table_id_dict)
# table_to_text("input.pdf")


